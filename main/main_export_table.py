# -*- coding: utf-8 -*-
import calendar
import os
import re
from pprint import pprint
import pandas as pd
from sqlalchemy import create_engine, text
from render_galderma.helper.logger_helper import LoggerSimple
from render_galderma.helper.reader_helper import load_json
from render_galderma.helper.data_table_helper import sort_price_segment
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Color
from openpyxl.styles import Font
import warnings

warnings.filterwarnings("ignore")
logger = LoggerSimple(name=__name__).logger
from constants import *

from render_galderma.input_template.serum import config_to_table_serum



config_dicts = {}
for i in [obj for name, obj in globals().items() if isinstance(obj, dict) and name.startswith("config_to_table")]:
    config_dicts.update(i)




def sort_multiple_columns(df):
    for column, sorted_cate in columns_and_categories.items():
        if column in df.columns:
            df['sort_order'] = df[column].apply(lambda x: sorted_cate.index(x) if x in sorted_cate else -1)
            if 'sale' in df.columns:
                df = df.sort_values(by=['sort_order', 'sale'], ascending=[True, False], na_position='last')
            if 'revenue' in df.columns:
                df = df.sort_values(by=['sort_order', 'revenue'], ascending=[True, False], na_position='last')
            df = df.drop(columns=['sort_order'])
    if 'price_range' in df.columns:
        df = sort_price_segment(df, 'price_range')
    return df


def get_location_table_in_template(file_template_path, table_config, sheet_name):
    table_key = table_config.get('key')
    row_add, col_add = table_config.get('row'), table_config.get('col')
    df = pd.read_excel(file_template_path, engine='openpyxl', header=None, sheet_name=sheet_name)

    for idx, row in df.iterrows():
        for idx_col, col in enumerate(row):
            if col == table_key:
                logger.info(f'{table_key} : row={idx}+{row_add} ; col={idx_col}+{col_add}')
                return idx, idx_col, idx + row_add, idx_col + col_add
    return None, None, None, None


def get_range_time(start_date, end_date):
    start_dt = datetime.strptime(start_date, "%Y%m")
    end_dt = datetime.strptime(end_date, "%Y%m")
    start_str = start_dt.strftime("01/%m/%Y")
    last_day = calendar.monthrange(end_dt.year, end_dt.month)[1]
    end_str = f"{last_day}/{end_dt.strftime('%m/%Y')}"
    return f"({start_str} - {end_str})"


range_time = get_range_time(start_date, end_date)


def write_table_to_excel(table_key, writer, sheet_name, table_config, config_template, df, no_col, no_row):
    worksheet = writer.sheets[sheet_name]
    df.fillna(0, inplace=True)
    if 'sale' in df.columns:
        df = df.sort_values(by='sale', ascending=False)
    if 'Sales Value' in df.columns:
        df = df.sort_values(by='Sales Value', ascending=False)
    if 'price_range' in df.columns:
        df = sort_price_segment(df, 'price_range')
    # if "overview" in table_key or "_maricosea_type" in table_key:
    #     df = sort_multiple_columns(df)
    header = True


    for column in df.columns:
        if column in [0, 1,2 ,3]:
            map_rows = {}
            for row_value in df[column]:
                map_rows[row_value] = row_value
            map_rows.update(config_template.get('row_alias'))
            df[column] = df[column].map(map_rows)

    map_columns = {}
    for column in df.columns:
        map_columns[column] = column
    map_columns.update(config_template.get('column_alias'))
    # if 'output_partner_brand' in df.columns or table_key == "tb_overview_brand":
    #     map_columns.update(config_template.get('column_min_alias'))
    df = df.rename(columns=map_columns)

    # Clearing contents by writing blank cells
    for row in range(no_row, no_row + df.shape[0] + 1):
        for col in range(no_col, df.shape[1] + 1):
            worksheet.write_blank(row, col, None)
    if table_config.get('is_transpose') is True:
        df = df.T.reset_index()
        header = False
    if table_config.get('is_hide_header') is True:
        header = False
    df.to_excel(
        writer, sheet_name=sheet_name,
        startrow=no_row, startcol=no_col, index=False,
        header=header
    )
    if ('.' not in table_key) or ('_' in table_key):
        cell_format = workbook.add_format()
        value_change = ''
        worksheet.write(source_idx_row, source_idx_col, value_change, cell_format)


def get_table_config(config_template, table_key):
    lst_table = config_template.get('lst_table')
    map_table = {i.get('key'): i for i in lst_table}
    if table_key in map_table:
        return map_table.get(table_key)
    if table_key == 'overview_total_market':
        return {
            'key': table_key,
            "row": 0,
            "col": 2
        }
    elif table_key.startswith('tp_us_'):
        return {
            'key': table_key,
            "row": 0,
            "col": 1,
            # 'is_transpose': True,
            'is_hide_header': True
        }
    elif table_key.startswith('tp_'):
        return {
            'key': table_key,
            "row": 0,
            "col": 1,
            'is_transpose': True
        }
    elif table_key.startswith('client_'):
        return {
            'key': table_key,
            "row": 0,
            "col": 2,
            'is_hide_header': True
        }
    return {
        'key': table_key,
        "row": 0,
        "col": 1
    }


now = datetime.now()
now_str = now.strftime("%Y%m%d_%H%M%S")
with pd.ExcelWriter(f"{file_output_to_excel_path}_{now_str}.xlsx", engine='xlsxwriter',
                    engine_kwargs={'options': {'strings_to_urls': False}}) as writer:
    workbook = writer.book
    dict_autocf = {}
    for sheet_name, dict_sub_config in config_dicts.items():
        if not dict_sub_config:
            workbook.add_worksheet(sheet_name)
            continue

        columns_and_categories = {
            "partner_function": ["Male", "Female", "Total"],
            "partner_type_shop": ["Shop Mall", "Shop Non-mall", "Total"],
            "partner_chanel": ["Shopee", "Lazada", "TikTok Shop", "Total"],
            "maricosea_share": ["Marico Sea Non-Official", "Marico Sea Official", "Total Marico Sea"],
            "category_maricosea": ["Tổng cộng", "Shampoo", "Hybrid (2in1, 3in1, 5in1)", "Hair styling", "Shower gel",
                                   "Deo",
                                   "Facewash", "Hair Conditioner", "Perfume", "Combo (Shampoo and Hair Conditioner)"]
        }
        print(f'sheet_name={sheet_name}')
        logger.info(dict_sub_config.keys())

        config_template = load_json(file_input_config_path)
        lst_table = config_template.get('lst_table')
        df = pd.read_excel(file_input_template_path, header=None, sheet_name=sheet_name)

        for table_key in dict_sub_config.keys():
            if table_key.startswith('infor_autocf_'):
                brand = table_key.split('_')[2]
                no1_autocf_share = f"autocf_share_{brand}_model_1"
                no1_autocf_platformbytime = f"autocf_platformbytime_{brand}_model_1"
                for idx, row in df.iterrows():
                    for idx_col, col in enumerate(row):
                        if col == no1_autocf_share:
                            # lưu lại toạ độ để copy format bảng autorender
                            dict_autocf.setdefault(sheet_name, {}).update(
                                {f"autocf_share_{brand}_model_1": [idx, idx_col + 1]}
                            )
                            for i in range(2, dict_sub_config.get(table_key)+1):
                                col_index = (idx_col + 15) * (i-1)
                                dict_autocf.setdefault(sheet_name,{}).update(
                                    {f"autocf_share_{brand}_model_{i}": [idx, col_index + 1]}
                                )
                        elif col == no1_autocf_platformbytime:
                            for i in range(2, dict_sub_config.get(table_key)+1):
                                col_index = (idx_col + 15) * (i-1)
                                dict_autocf.setdefault(sheet_name,{}).update(
                                    {f"autocf_platformbytime_{brand}_model_{i}": [idx, col_index + 1]}
                                )

        df.to_excel(writer, index=False, header=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        # worksheet.write("C3", range_time)
        for idx, col in enumerate(df):
            series = df[col]
            max_len = max((
                series.astype(str).map(len).max(),
                len(str(series.name))
            )) + 1
            worksheet.set_column(idx, idx, max_len)
        for idx, table_key in enumerate(dict_sub_config.keys()):
            if table_key.startswith('infor_autocf_'):
                continue
            logger.info(f'{idx} - prepare table: table_key={table_key}')
            if table_key is None:
                continue
            table_config = get_table_config(config_template, table_key)
            data = dict_sub_config.get(table_key)
            if data.empty:
                logger.info(f"dataframe not found for {table_config.get('key')}")
                continue
            if table_config is None:
                logger.info(f"table config not found for {table_key}")
                continue
            if table_key.startswith('autocf') and not table_key.endswith('_1'):
                row = dict_autocf.get(sheet_name).get(table_key)[0]
                col = dict_autocf.get(sheet_name).get(table_key)[1]

            else:
                source_idx_row, source_idx_col, row, col = get_location_table_in_template(
                    file_template_path=file_input_template_path,
                    table_config=table_config,
                    sheet_name=sheet_name
                )
                if source_idx_row is None or source_idx_col is None or row is None or col is None:
                    logger.info(f"location not found for {table_key}")
                    continue
            if table_key.startswith('autocf_share'):
                table_config = {'is_hide_header': True}


            write_table_to_excel(table_key, writer, sheet_name=sheet_name, df=data,
                                 table_config=table_config,
                                 config_template=config_template,
                                 no_col=col, no_row=row)

        print(f'done sheet: {sheet_name}')


def copy_format_rectangle(ws_source, ws_target, row_idx, col_idx, num_rows, num_cols, row_idx_target, col_idx_target):
    for i in range(num_rows):
        for j in range(num_cols):
            src_cell = ws_source.cell(row=row_idx + i, column=col_idx + j)
            target_cell = ws_target.cell(row=row_idx_target + i, column=col_idx_target + j)
            if src_cell.value and isinstance(src_cell.value, str) and 'TOTAL' in src_cell.value:
                target_cell.value = src_cell.value
            if src_cell.data_type == 'f':
                target_cell.value = src_cell.value
            if src_cell.has_style:
                target_cell.font = src_cell.font.copy()
                target_cell.number_format = src_cell.number_format
                target_cell.alignment = src_cell.alignment.copy()
                target_cell.fill = src_cell.fill.copy()
                def copy_side(side):
                    if side.color:
                        try:
                            rgb_value = str(side.color.rgb) if side.color.rgb else None
                        except AttributeError:
                            rgb_value = None
                        if rgb_value and len(rgb_value) == 6:
                            color = Color(rgb="FF" + rgb_value)
                        elif rgb_value and len(rgb_value) == 8:
                            color = Color(rgb=rgb_value)
                        else:
                            color = None
                    else:
                        color = None
                    return Side(style=side.style, color=color)

                target_cell.border = Border(
                    left=copy_side(src_cell.border.left),
                    right=copy_side(src_cell.border.right),
                    top=copy_side(src_cell.border.top),
                    bottom=copy_side(src_cell.border.bottom)
                )
                if target_cell.fill is None or target_cell.fill.fill_type is None:
                    target_cell.fill = white_fill
                else:
                    if isinstance(target_cell.fill, PatternFill) and target_cell.fill.start_color.type == 'theme':
                        target_cell.fill = white_fill
                    elif target_cell.fill.start_color.index == '00000000':
                        target_cell.fill = white_fill
    for merge_range in ws_source.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merge_range.min_row, merge_range.min_col, merge_range.max_row, merge_range.max_col

        if row_idx <= min_row <= row_idx + num_rows and col_idx <= min_col <= col_idx + num_cols:
            new_min_row = row_idx_target + (min_row - row_idx)
            new_max_row = row_idx_target + (max_row - row_idx)
            new_min_col = col_idx_target + (min_col - col_idx)
            new_max_col = col_idx_target + (max_col - col_idx)
            ws_target.merge_cells(start_row=new_min_row, start_column=new_min_col,
                                  end_row=new_max_row, end_column=new_max_col)


white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
sheets_in_template = pd.ExcelFile(file_input_template_path).sheet_names
wb_source = load_workbook(file_input_template_path)
wb_target = load_workbook(f"{file_output_to_excel_path}_{now_str}.xlsx")
for sheet_name in sheets_in_template:
    if sheet_name in config_dicts:
        print(f"Start copy format sheet: {sheet_name}")
        ws_source = wb_source[sheet_name]
        ws_target = wb_target[sheet_name]
        if config_dicts.get(sheet_name):
            for col, col_dim in ws_target.column_dimensions.items():
                ws_target.column_dimensions[col].width = 24
        else:
            for col in ws_source.column_dimensions:
                ws_target.column_dimensions[col].width = ws_source.column_dimensions[col].width

        for row_idx, row_dim in ws_source.row_dimensions.items():
            ws_target.row_dimensions[row_idx].height = row_dim.height

        for row in ws_source.iter_rows():
            for cell in row:
                target_cell = ws_target[cell.coordinate]
                if not config_dicts.get(sheet_name):
                    target_cell.value = cell.value
                if cell.data_type == 'f':
                    target_cell.value = cell.value
                if cell.has_style:
                    target_cell.font = cell.font.copy()
                    target_cell.number_format = cell.number_format
                    target_cell.alignment = cell.alignment.copy()
                    target_cell.fill = cell.fill.copy()


                    def copy_side(side):
                        """
                        Sao chép đối tượng Side và điều chỉnh màu nếu cần.
                        """
                        if side.color:
                            try:
                                rgb_value = str(side.color.rgb) if side.color.rgb else None
                            except AttributeError:
                                rgb_value = None
                            if rgb_value and len(rgb_value) == 6:
                                color = Color(rgb="FF" + rgb_value)
                            elif rgb_value and len(rgb_value) == 8:
                                color = Color(rgb=rgb_value)
                            else:
                                color = None
                        else:
                            color = None
                        return Side(style=side.style, color=color)


                    target_cell.border = Border(
                        left=copy_side(cell.border.left),
                        right=copy_side(cell.border.right),
                        top=copy_side(cell.border.top),
                        bottom=copy_side(cell.border.bottom)
                    )
                    if target_cell.fill is None or target_cell.fill.fill_type is None:
                        target_cell.fill = white_fill
                    else:
                        if isinstance(target_cell.fill, PatternFill) and target_cell.fill.start_color.type == 'theme':
                            target_cell.fill = white_fill
                        elif target_cell.fill.start_color.index == '00000000':
                            target_cell.fill = white_fill
                # if target_cell.value == 'svr':
                #     current_font = target_cell.font
                #     target_cell.font = Font(
                #         name=current_font.name,
                #         size=current_font.size,
                #         bold=True,
                #         italic=current_font.italic,
                #         underline=current_font.underline,
                #         color=current_font.color
                #     )
        for merge_range in ws_source.merged_cells.ranges:
            ws_target.merge_cells(str(merge_range))
        ws_target.sheet_view.showGridLines = False
        config_copy_format_auto = dict_autocf.get(sheet_name)
        row_sr = config_copy_format_auto.get('autocf_share_loreal_model_1')[0] - 1
        col_sr = config_copy_format_auto.get('autocf_share_loreal_model_1')[1]
        # for table_key in config_copy_format_auto:
        #     if table_key.startswith("autocf_share_"):
        #         row = config_copy_format_auto.get(table_key)[0] - 1
        #         col = config_copy_format_auto.get(table_key)[1]
        #         copy_format_rectangle(ws_source,ws_target, row_sr, col_sr,8,15,row +2, col)
        print(f"Done copy format sheet: {sheet_name}")
wb_target.save(f"{file_output_to_excel_path}_{now_str}.xlsx")
