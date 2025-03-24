from openpyxl import load_workbook
from render_galderma.main.constants import *


def extract_keys_from_excel(file_path):
    wb = load_workbook(filename=file_path,
                       data_only=True)
    extracted_keys = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value,
                                             str) and "_" in cell.value:
                    extracted_keys.setdefault(sheet_name, {}).update({cell.value: f"df_{cell.value}"})
    return extracted_keys


file_path = file_input_template_path
dict_config = extract_keys_from_excel(file_path)
print(dict_config)

dict = {'Serum': {'overview_total_market': 'df_overview_total_market', 'overview_channel': 'df_overview_channel',
                  'overview_pricerange': 'df_overview_pricerange', 'overview_function': 'df_overview_function',
                  'overview_ingredient': 'df_overview_ingredient', 'tp_b_1_1': 'df_tp_b_1_1', 'tp_b_1_2': 'df_tp_b_1_2',
                  'tp_b_1_3': 'df_tp_b_1_3', 'tp_b_1_4': 'df_tp_b_1_4', 'tp_b_1_5': 'df_tp_b_1_5',
                  'brand_b_1_1': 'df_brand_b_1_1', 'brand_b_1_2': 'df_brand_b_1_2', 'brand_b_1_3': 'df_brand_b_1_3',
                  'brand_b_1_4': 'df_brand_b_1_4', 'brand_b_1_5': 'df_brand_b_1_5', 'client_b_1_1': 'df_client_b_1_1',
                  'client_b_1_2': 'df_client_b_1_2', 'client_b_1_3': 'df_client_b_1_3',
                  'client_b_1_4': 'df_client_b_1_4', 'client_b_1_5': 'df_client_b_1_5', 'ANTI_AGING': 'df_ANTI_AGING',
                  'ANTI_ACNE': 'df_ANTI_ACNE', 'tp_c_1_1': 'df_tp_c_1_1', 'tp_c_1_2': 'df_tp_c_1_2',
                  'tp_c_1_3': 'df_tp_c_1_3', 'tp_c_1_4': 'df_tp_c_1_4', 'tp_c_1_5': 'df_tp_c_1_5',
                  'tp_c_1_6': 'df_tp_c_1_6', 'brand_c_1_1': 'df_brand_c_1_1', 'brand_c_1_2': 'df_brand_c_1_2',
                  'brand_c_1_3': 'df_brand_c_1_3', 'brand_c_1_4': 'df_brand_c_1_4', 'brand_c_1_5': 'df_brand_c_1_5',
                  'brand_c_1_6': 'df_brand_c_1_6', 'client_c_1_1': 'df_client_c_1_1', 'client_c_1_2': 'df_client_c_1_2',
                  'client_c_1_3': 'df_client_c_1_3', 'client_c_1_4': 'df_client_c_1_4',
                  'client_c_1_5': 'df_client_c_1_5', 'client_c_1_6': 'df_client_c_1_6',
                  'tp_us_loreal': 'df_tp_us_loreal', 'autocf_share_loreal_model_1': 'df_autocf_share_loreal_model_1',
                  'loreal_platform': 'df_loreal_platform',
                  'autocf_platformbytime_loreal_model_1': 'df_autocf_platformbytime_loreal_model_1',
                  'tp_us_laroche': 'df_tp_us_laroche',
                  'autocf_share_laroche_model_1': 'df_autocf_share_laroche_model_1',
                  'laroche_platform': 'df_laroche_platform',
                  'autocf_platformbytime_laroche_model_1': 'df_autocf_platformbytime_laroche_model_1',
                  'tp_us_garnier': 'df_tp_us_garnier',
                  'autocf_share_garnier_model_1': 'df_autocf_share_garnier_model_1',
                  'garnier_platform': 'df_garnier_platform',
                  'autocf_platformbytime_garnier_model_1': 'df_autocf_platformbytime_garnier_model_1'}}
